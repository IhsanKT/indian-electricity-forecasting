"""Turn 29 heterogeneous Mendeley workbooks into one clean hourly demand series.

The source mixes three formats (see docs/data_audit.md section 2):

  regime A  Sep 2021 - Sep 2022   5-minute SCADA export
  regime B  Oct 2022 - Dec 2023   10-second SCADA export
  regime C  Jan 2024 - Jun 2025   already aggregated to hourly by the depositor

Regimes A and B are aggregated here by **hourly mean**, which the audit established is the
same rule the depositor used for regime C. Using a different rule would put a ~1.5%
structural break at 2024-01-01, in the middle of the train/test span.

Every gap fill uses past information only. Interpolating across a gap with later values
would leak the future into the training data.
"""
from __future__ import annotations

import json
import re
import sys
import time
from typing import Iterator

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from src import config


# --------------------------------------------------------------------------------------
# Reading
# --------------------------------------------------------------------------------------
def _sampling_interval_seconds(path) -> float | None:
    """Read the native sample interval out of the hidden `_osi_config` sheet.

    The sheet stores JSON like {"r":300.0,...} where r is seconds between samples. Used to
    verify the regime rather than inferring it from file size.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "_osi_config" not in wb.sheetnames:
            return None
        ws = wb["_osi_config"]
        for row in ws.iter_rows(max_row=2, max_col=1, values_only=True):
            cell = row[0]
            if isinstance(cell, str) and '"r"' in cell:
                m = re.search(r'"r"\s*:\s*([0-9.]+)', cell)
                if m:
                    return float(m.group(1))
        return None
    finally:
        wb.close()


def _read_scada_workbook(path) -> pd.Series:
    """Stream Time + NLDC_DEMAND|P out of a regime A/B SCADA workbook."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[config.SCADA_SHEET]
        stamps: list = []
        values: list = []
        for row in ws.iter_rows(min_row=config.SCADA_HEADER_ROWS + 1, values_only=True):
            ts = row[config.SCADA_TIME_COL]
            if ts is None:
                continue
            stamps.append(ts)
            values.append(row[config.SCADA_DEMAND_COL])
    finally:
        wb.close()
    return pd.Series(
        pd.to_numeric(pd.Series(values), errors="coerce").to_numpy(),
        index=pd.to_datetime(pd.Series(stamps)),
        name="demand",
    )


def _read_hourly_workbook(path) -> pd.Series:
    """Read the regime C workbook.

    Two traps here, both from docs/data_audit.md section 6: the sheet carries six trailing
    Excel summary rows (Minimum/Maximum/Average/Sum) which yield a 2.5-billion-MW maximum
    if kept, and Timestamp is a DD-MM-YYYY *string*, so dayfirst is mandatory.
    """
    df = pd.read_excel(path, sheet_name=config.HOURLY_SHEET)
    ts = pd.to_datetime(df["Timestamp"], dayfirst=True, errors="coerce")
    demand = pd.to_numeric(df["Demand (MW)"], errors="coerce")
    keep = ts.notna()
    dropped = int((~keep).sum())
    if dropped:
        print(f"      dropped {dropped} non-timestamp rows (Excel summary block)")
    return pd.Series(demand[keep].to_numpy(), index=pd.DatetimeIndex(ts[keep]), name="demand")


# --------------------------------------------------------------------------------------
# Cleaning
# --------------------------------------------------------------------------------------
def mask_invalid(s: pd.Series) -> tuple[pd.Series, dict[str, int]]:
    """Replace physically impossible readings with NaN, before any aggregation.

    SCADA telemetry drops to exactly 0.0 MW on sensor failure. Those are not demand, and
    averaging them into an hourly bucket drags it down by up to 2.3%.
    """
    zeros = int((s == 0).sum())
    out = s.replace(0.0, np.nan)
    implausible = int(((out < config.PLAUSIBLE_MIN_MW) | (out > config.PLAUSIBLE_MAX_MW)).sum())
    out = out.where(
        (out >= config.PLAUSIBLE_MIN_MW) & (out <= config.PLAUSIBLE_MAX_MW)
    )
    return out, {"zero_dropouts": zeros, "implausible": implausible}


def to_hourly(s: pd.Series) -> pd.Series:
    """Aggregate a sub-hourly series to hourly means (the depositor's rule)."""
    return s.resample("h").mean()


# --------------------------------------------------------------------------------------
# Pipeline
# --------------------------------------------------------------------------------------
def _iter_workbooks() -> Iterator:
    yield from sorted(config.MENDELEY_RAW_DIR.glob("*.xlsx"))


def load_all_hourly(verbose: bool = True) -> tuple[pd.Series, list[dict]]:
    """Read every workbook and return one concatenated hourly series plus per-file stats.

    Each workbook covers a whole calendar month, so hourly buckets never straddle files and
    per-file aggregation is safe.
    """
    pieces: list[pd.Series] = []
    stats: list[dict] = []
    for path in _iter_workbooks():
        t0 = time.time()
        interval = _sampling_interval_seconds(path)
        is_hourly_file = path.name == config.HOURLY_FILE_NAME
        if is_hourly_file:
            raw = _read_hourly_workbook(path)
            regime = "C (pre-aggregated hourly)"
        else:
            raw = _read_scada_workbook(path)
            regime = f"{'A' if interval == 300.0 else 'B'} (SCADA {interval:g}s)"

        n_raw = len(raw)
        dupes = int(raw.index.duplicated().sum())
        if dupes:
            raw = raw[~raw.index.duplicated(keep="first")]
        raw = raw.sort_index()

        cleaned, flags = mask_invalid(raw)
        hourly = cleaned if is_hourly_file else to_hourly(cleaned)

        pieces.append(hourly)
        stats.append({
            "file": path.name,
            "regime": regime,
            "native_interval_s": interval,
            "raw_rows": n_raw,
            "duplicate_timestamps": dupes,
            "hourly_rows": int(len(hourly)),
            "nan_hours": int(hourly.isna().sum()),
            **flags,
            "start": str(hourly.index.min()),
            "end": str(hourly.index.max()),
        })
        if verbose:
            print(f"  {path.name:<32} {regime:<22} {n_raw:>7,} -> {len(hourly):>5,} h "
                  f"({time.time()-t0:.1f}s)", flush=True)

    series = pd.concat(pieces).sort_index()
    series = series[~series.index.duplicated(keep="first")]
    return series, stats


def build_processed(verbose: bool = True) -> pd.Series:
    """Full Phase 1 pipeline. Writes the processed series and a preprocessing report."""
    print("Reading workbooks ...")
    series, per_file = load_all_hourly(verbose=verbose)

    # Reindex onto a strict hourly grid so gaps become explicit NaNs rather than silence.
    full_index = pd.date_range(series.index.min(), series.index.max(), freq="h")
    before = len(series)
    series = series.reindex(full_index)
    missing_idx = series.index[series.isna()]

    print(f"\nGrid: {full_index[0]} -> {full_index[-1]}  ({len(full_index):,} hours)")
    print(f"  observed hours : {before:,}")
    print(f"  missing hours  : {len(missing_idx):,} "
          f"({len(missing_idx)/len(full_index)*100:.3f}%)")
    if len(missing_idx):
        by_month = pd.Series(1, index=missing_idx).resample("MS").sum()
        print("  missing by month (non-zero):")
        print(by_month[by_month > 0].to_string())

    # PAST-ONLY fill. Forward fill carries the last observed value forward; it never looks
    # ahead. Interpolation would average across the gap using future values -> leakage.
    filled = series.ffill()
    # A leading NaN cannot be filled from the past; drop rather than back-fill.
    n_leading = int(filled.isna().sum())
    if n_leading:
        filled = filled.iloc[n_leading:]
        print(f"  dropped {n_leading} leading hours with no prior observation")

    filled = filled.tz_localize(config.TIMEZONE)

    out = filled.rename("demand").to_frame()
    out.index.name = "timestamp"
    config.PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    out.to_parquet(config.PROCESSED_SERIES_PATH)

    report = {
        "source_doi": config.MENDELEY_DOI,
        "timezone": config.TIMEZONE,
        "aggregation": "hourly mean (matches depositor rule for regime C; see data_audit.md s.7)",
        "gap_fill": "forward fill (past information only)",
        "grid_start": str(full_index[0]),
        "grid_end": str(full_index[-1]),
        "grid_hours": int(len(full_index)),
        "observed_hours": int(before),
        "missing_hours_filled": int(len(missing_idx)),
        "missing_hours_list": [str(t) for t in missing_idx[:200]],
        "final_rows": int(len(filled)),
        "per_file": per_file,
    }
    config.GAP_REPORT_PATH.write_text(json.dumps(report, indent=2), encoding="utf-8")

    print(f"\nWrote {config.PROCESSED_SERIES_PATH}  ({len(filled):,} hourly rows)")
    print(f"Wrote {config.GAP_REPORT_PATH}")
    print(f"\ndemand MW: min={filled.min():,.1f}  mean={filled.mean():,.1f}  max={filled.max():,.1f}")
    print(f"peak at {filled.idxmax()}")
    return filled


def load_processed() -> pd.Series:
    """Load the processed hourly series (Phase 2+ entry point)."""
    df = pd.read_parquet(config.PROCESSED_SERIES_PATH)
    return df["demand"]


if __name__ == "__main__":
    build_processed()
    sys.exit(0)
